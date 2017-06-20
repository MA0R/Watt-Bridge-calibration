"""
The main thread for the Watt bridge calibration, all structure is there but
it needs all sorts of functions set up to properly control the instruments.
All instruments except the fluke are a very simple class which allows
them to be controlled with the self.com function, this wraps each instrument
communication with a check, to see if the command was sucesful or not.
The fluke instrument is not one such class, it is the only exception.

Improvements:
Turn instruments into more specialised classes, so as to reuse code
and remove some info from this thread. For example, a more specific watt bridge
instrument that has a function for setting dials.
Lots of the communications can be reduced this way and then the classes can be reused
later.
"""
import stuff
import time
import numpy as np
import inst
import openpyxl
import threading
import swerlein
import ReadExcel_v3_maor

class Thread(threading.Thread):
    """
    Watt bridge main thread
    """
    def __init__(self, notify_window, EVT, param, data):
        """ Initialisation requires the param and data to be sent.
        Param holds the instruments, data is the thread-safe data storage place."""
        threading.Thread.__init__(self, notify_window, EVT, param)
        self._want_abort = 0
        self.MadeSafe = False #Did we mak_safe flag.
        self.printing_status = True #if it is false, do not PrintSave.
        self.inst_bus = param['inst_bus'] #visa or simulated visa

        #Open the main excel file, and setup the sheet from which instructions are read.
        self.wb = openpyxl.load_workbook(param['wb'],data_only = True)
        self.data_sh = self.wb['Data'] #First data sheet, contains measurement data.
        self.constants_reader = ReadExcel_v3_maor.WATTDATA(param['wb'])
        
        #collect the necessary instruments
        self.rd = param['rd']
        self.rd.create_instrument()
        self.wbridge = param['wbridge']
        self.wbridge.create_instrument()
        self.meter = param['meter']
        self.meter.create_instrument()
        self.count = param['count']
        self.count.create_instrument()
        self.fluke = param['fluke']
        self.ch5500 = param['ch5500']
        self.ch5500.create_instrument()
        self.ch5050 = param['ch5050']
        self.ch5050.create_instrument()
        self.heg = param['heg']
        self.heg.create_instrument()

        #Columns that are used multiple times:
        self.source_col = 2

        #Source channels need to be decided at the start:
        self.source_c3 = raw_input("Source channel 3 (0/1): ")
        self.source_c2 = raw_input("Source channel 2 (0/1): ")
        self.source_c1 = raw_input("Source channel 1 (0/1): ")

        #Create the log file names as a time/date stamp of the init.
        time_bit = time.strftime("%Y.%m.%d.%H.%M.%S, ",time.localtime())
        log_file_name = 'log.'+time_bit+".txt"
        self.raw_file_name = 'raw.'+time_bit
        self.logfile = open(log_file_name,'w')
        #It's all downhill from here!
        self.start()
        
    def abort(self):
        """abort worker thread."""
        # Method for use by main thread to signal an abort
        self._want_abort = 1
        
    def PrintSave(self, text):
        """
        Prints a string, and saves to the log file too.
        """
        if self.printing_status == True:
            if self._want_abort:
                #stop the pointless printing if the thread wants to abort
                return
            else:
                self.logfile.write(str(text)+"\n")
                print(repr(text))
    def com(self, command,send_item=None):
        """
        Operate "command" on "send_item" if it is specified. Command is
        a method of the instrument class.
        All commands [Fluke? RD31?] go through this function, as it wraps
        each command
        with a check to see if the thread should abort.
        Recieve the status of the command from the instrument
        class, and if it failes this will flag the thread to abort.
        """
        #print(send_item)
        if not self._want_abort:
            if send_item != None:
                sucess,val,string = command(send_item)
                if sucess == False:
                    self.abort()
                self.PrintSave(string)
                return val
            else:
                sucess,val,string = command()
                if sucess == False:
                    self.abort()
                self.PrintSave(string)
                return val
        else:
            return 0
        
    def MakeSafe(self):
        #probably not needed for internal use?
        pass
        
    def read_grid_cell(self,row,col,grid = None):
        """ Read a cell in an openpyxl worksheet object.
        If no sheet is specified, defaults to the data sheet."""
        value = 0
        if grid == None:
            grid = self.data_sh
            
        if not self._want_abort:
            value = grid.cell(row=row, column=col).value
        return value
        
    def end(self):
        """
        clean up, turn instruments off [make safe?]
        """
        self.PrintSave("##   READINGS COMPLETE   ##")

    def wait(self,wait_time):
        """
        safely loop until time is up, checking for abort
        """
        self.PrintSave("Waiting for {} seconds".format(wait_time))
        t=time.time()
        while time.time()<t+wait_time:
            if self._want_abort: #has the stop button been pushed?
                break
        return

    def initialise_counter(self):
        #only writing init for AG53230A
        self.com(self.count.set_value,"*rst;*cls;*sre 0;ese 0;stat:pres;:INP1:FILT:LPASS:STAT 1\n")
        self.com(self.count.set_value,"sens:freq:gate:sour time\n")
        self.com(self.count.set_value,":inp1:rang 50\n")
        self.com(self.count.set_value,":sens:rosc:sour int\n")
        if channel1_filter==True:
            self.com(self.count.set_value,";:INP1:FILT:LPAS:STAT 1\n")
        else:
            self.com(self.count.set_value,";:inp1:filt:lpas:stat 0\n")
        if ch_1_trig_level==0:
            self.com(self.count.set_value,":inp1:coup dc;:inp1:slope pos;:inp1:lev:abs 3V;\n")
        else:
            self.com(self.count.set_value,":inp1:coup dc;:inp1:slope pos;:inp1:lev:abs 6V;\n")
        if ch2_trig_level==0:
            self.com(self.count.set_value,":inp2:coup dc;:inp2:slope pos;:inp2:lev:abs 3V;\n")
        else:
            self.com(self.count.set_value,":inp2:coup dc;:inp2:slope pos;:inp2:lev:abs 6V;\n")
            
    def initialise_radian(self,row):
        """
        Initialisation process as in TP, with the phase sent
        to it after the wait
        """
        rd_phase_col = 16
        watts_or_vars_cel = 13
        
        self.wait(1)
        rd_phase = self.read_grid_cell(row,rd_phase_col)
        if rd_phase != None:
            self.com(self.rd.set_value,rd_phase)
        watts_or_vars = self.read_grid_cell(row,watts_or_vars_cel)
        if watts_or_vars != None:
            self.set_rd_output_pulse('prog_radian_id',1,1,rd_phase)
        else:
            pass
            #  No idea what goes on here :(
            
    def set_rd_output_pulse(self, a, b, c, d):
        pass

    def set_power(self, row):
        divider_range = self.read_grid_cell(row,6)
        shunt = self.read_grid_cell(row,7)
        ct_ratio = self.read_grid_cell(row,8)
        heg_freq = self.read_grid_cell(row,9)

        #This part cmmunicates with the WB
        #Need to set "baud rate=9600, parity=N, bits=8, stop bits=1?
        self.com(self.wbridge.set_value,"DV\rDV\rW0000\rV0000\rA01\rB01\r")
        #It is possible that these will have to be sent as seperate commands.
        #The test point code wait for the machine to respond after each command
        #But visa does something similar too, raising an exception otherwise?
        if divider_range == '60':
            self.com(self.wbridge,"R060\r")
        else:
            self.com(self.wbridge.set_value,"R{}\r".format(divider_range))
        self.com(self.wbridge.set_value,"WP-")
        self.com(self.wbridge.set_value,"VP-")
        self.wait(0.5)
    
    def power_sources(self,row):
        """Read which source type we have, and power the correct one"""
        source_type = self.read_grid_cell(row,self.source_col)
        if source_type == 'FLUKE':
            self.power_flue(row)
        elif source_type == 'FLUHIGH':
            self.power_fluke(row)
        elif source_type == 'CH':
            self.power_ch(row)
        elif source_type == 'HEG':
            self.power_heg(row)
        else:
            #abort since there is no valid source type
            self.PrintSave("Invalid source '{}' at row {}"\
                           .format(source_type,row))
            self._want_abort = 1

    def power_fluke(self,row):
        set_volts = self.read_grid_cell(row,4)
        set_phase = self.read_grid_cell(row,5)
        set_amps = self.read_grid_cell(row,3)
        set_freq = self.read_grid_cell(row,9)
        self.fluke.initialise()#can have fluke ramp time sent here
        #   AND WHAT ELSE?
        
        
    def power_ch(self,row):
        """CH startup routine"""
        #repeated code? Already retrieved these cell values before.
        #could simplify
        set_volts = self.read_grid_cell(row,4)
        set_phase = self.read_grid_cell(row,5)
        set_amps = self.read_grid_cell(row,3)
        set_freq = self.read_grid_cell(row,9)
        #check first to see if ch5050 is wired in.
        if self.ch5050:
            self.com(ch5050,"SI\n")
        self.com(self.ch5500.set_value,"S\n")
        if self.HVamp:
            self.com(self.ch5500.set_value,"O0\n")
            set_volts = self.amp_in_volts(set_volts)
        else:
            self.com(self.ch5500.set_value,"O180\n")
        self.com(self.ch5500.set_value,"R{}\n".format(set_volts))
        change_offset = True
        if change_offset:
            self.com(self.ch5500.set_value,"O0\n")
        self.com(self.ch5500.set_value,"V{}\n".format(np.abs(set_amps)))
        self.com(self.ch5500.set_value,"P{}F{}N\n".format(set_phase,set_freq))
        if self.ch5050:
            self.com(self.ch5050.set_value,"N\n")
            self.com(self.ch5500.set_value,"N\n")
            self.com(self.ch5050.set_value,"O\n")
        self.wait(6)#(60)
        
    def power_heg(self,row):
        """Power the PL10 or HEG"""
        #PL10 in test-point code
        set_volts = self.read_grid_cell(row,4)
        set_amps = self.read_grid_cell(row,3)
        set_phase = self.read_grid_cell(row,5)
        
        self.com(self.heg.set_value,":SOUR:SCOND 4\n")
        self.com(self.heg.set_value,":SOUR:OPER:CURR3 {}\n".format(self.source_c3))
        self.com(self.heg.set_value,":SOUR:OPER:CURR2 {}\n".format(self.source_c2))
        self.com(self.heg.set_value,":SOUR:OPER:CURR1 {}\n".format(self.source_c1))
        self.com(self.heg.set_value,":SOUR:SOUR:MEN\n")
        #self.com(self.heg.set_value,":SOUR:SOUR:FREQ {}".format(heg_freq))
        #get set volts cell
        self.wait(1)
        self.com(self.heg.set_value,":SOUR:SOUR:VOLT1 {}\n".format(set_volts))
        #get set amps cell
        self.com(self.heg.set_value,":SOUR:SOUR:CURR1 {}\n".format(set_amps))
        #GET SET PHASE CELL
        self.wait(1)
        self.com(self.heg.set_value,":SOUR:SOUR:PHASE {}\n".format(set_phase))
        self.wait(2)
        self.com(self.heg.set_value,":SOUR:OPER:RUN\N")
        self.wait(6)#(60)
        
    def find_dial_settings(self,row):
        """
        First version: Simply replace the macro functions with python functions.
        Does not return anything, sends all control comamdns to the watt bridge from here.
        """
        set_volts = float(self.read_grid_cell(row,4))
        freq = self.frequency(set_volts)
        #Run the fft to obtain stuff
        fft_ref_v,fft_ref_p = self.fft_func(row,freq['uncal'])
        self.com(self.wbridge.set_value,'DD\r')
        #Now do the fft again
        fft_det_v,fft_det_p = self.fft_func(row,freq['uncal'])
        self.com(self.wbridge.set_value,'DV\r')

        creader = self.constants_reader #for simplicity
        sheet = self.wb['constants']

        ref_volts_v1 = fft_ref_v
        w_dial_alpha = 0
        watts_buffer_gain = creader.get_watts_buffer_gain('WP-',sheet)
        vars_dial_beta = 0
        vars_buffer_gain = creader.get_watts_buffer_gain('VP-',sheet)
        detector_amp_volts = fft_det_v
        detector_phase = (fft_det_p-fft_ref_p)*np.pi/180.0
        detector_amp_gain = creader.get_amp_gain('Min',sheet)
        freq['uncal'] = 51.3498573 #some random number that is reasonable, REMOVE THIS!
        c = [ref_volts_v1,w_dial_alpha,watts_buffer_gain,vars_dial_beta,vars_buffer_gain,\
             detector_amp_volts,detector_phase,detector_amp_gain]
        
        vals = self.calculate(c,freq,set_volts,sheet,row)
        #calcualte and send info. Sedning the info is actually part of 'refine dial settings'
        w_count = min(np.abs(vals['AlphaCount']),1023.)
        w_sign = (lambda x: 'WP-' if x>0 else 'WP+')(vals['Watts'])
        v_count = min(np.abs(vals['BetaCount']),1023.)
        v_sign = (lambda x: 'WP-' if x>0 else 'WP+')(vals['Vars'])
        self.load_dial_settings([w_count,w_sign,v_count,v_sign])
        
        #refine the dial settings: first re run fft.
        fft_det_v,fft_det_p = self.fft_func(row,freq['uncal'])

        #Now use the previously found settings to update all the calculations:
        w_dial_alpha = w_count/1024.
        watts_buffer_gain = creader.get_watts_buffer_gain(w_sign,sheet)
        vars_buffer_gain = creader.get_watts_buffer_gain(v_sign,sheet)
        vars_dial_beta = v_count/1024.
        c = [ref_volts_v1,w_dial_alpha,watts_buffer_gain,vars_dial_beta,vars_buffer_gain,\
             detector_amp_volts,detector_phase,detector_amp_gain]
        #recalculate the values:
        vals = self.calculate(c,freq,set_volts,sheet,row)

        #calculate and send the new settings
        w_count = min(np.abs(vals['AlphaCount']),1023.)
        w_sign = (lambda x: 'WP-' if x>0 else 'WP+')(vals['Watts'])
        v_count = min(np.abs(vals['BetaCount']),1023.)
        v_sign = (lambda x: 'WP-' if x>0 else 'WP+')(vals['Vars'])
        self.load_dial_settings([w_count,w_sign,v_count,v_sign])
    
    def frequency(self,volt_range):
        """Returns a dictionary of calibrated and uncalibrated frequencies.
        Could be specific to DVM later."""
        self.com(self.meter.set_value,'RESET;DCV {}\n'.format(volt_range))
        self.com(self.meter.set_value,'TARM HOLD;LFILTER ON;LEVEL 0,DC;FSOURCE ACDCV\n')
        self.com(self.meter.set_value,'FREQ\n')
        self.com(self.meter.set_value,'CAL? 245\n')
        freq_cal_const = float(self.com(self.meter.read))
        self.com(self.meter.set_value,'TARM SGL\n')
        self.wait(0.2)
        cal_freq = float(self.com(self.meter.read))
        uncal_freq = cal_freq/freq_cal_const
        return {'uncal':uncal_freq,'cal':cal_freq}
    
    def calculate_old_but_working(self,c,freq,set_volts,sheet,row):#As in the excel macros
        """I dont know what this code does, I copied the macros, and put all functions in one.
        Now will update the functions to use the complex maths in python, it will be much simpler.
        """
        creader = self.constants_reader
        #The constants are in the same order as those in the macros.
        Kraw = creader.get_Kraw(set_volts,sheet) #An array of three things, mean, uncertainty, DoF?
        #The excel sheet uses the uncal frequency, but i think it should be the calibrated frequency?
        Divider_Re = Kraw[0]
        Divider_Im = -Kraw[0]*(freq['uncal']*Kraw[1]+Kraw[2])
        RefVolts = c[0]
        shunt = self.read_grid_cell(row,7) #not specifying a sheet defaults to the data sheet
        shunt_impedance = creader.get_shuntraw(shunt,sheet)
        Shunt_Re = shunt_impedance[0]
        Shunt_Im = freq['uncal']*2*np.pi*shunt_impedance[1]*0.000001
        alpha = c[1]
        A = c[2]
        Re_R2R1 = creader.get_r2_by_r1(sheet)
        Im_R2R1 = creader.get_tc_r2_r1(sheet)*2*np.pi*freq['uncal']
        Beta = c[3]
        B = c[4]
        CondC1 = creader.get_g1(sheet)
        AdmtceC1 = creader.get_c1(sheet)*2*np.pi*0.000001*freq['uncal']
        r2_impedance = creader.get_r2(freq['uncal'], sheet)
        Re_R2 = r2_impedance.real
        Im_R2 = r2_impedance.imag
        U = c[5]
        DetPhse = c[6]
        G = c[7]
        DetAmpPhse = creader.get_det_amp_phase_gj(int(round(freq['uncal'],0)),sheet)
        Detcond = creader.get_gd(sheet)
        DetAdmtce = creader.get_yd(sheet)
        
        ScaleFactor_Re = -(Divider_Re * Shunt_Re + Shunt_Im * Divider_Im) * RefVolts * RefVolts / (Shunt_Re * Shunt_Re + Shunt_Im * Shunt_Im)
        ScaleFactor_Im = -(Divider_Im * Shunt_Re - Divider_Re * Shunt_Im) * RefVolts * RefVolts / (Shunt_Re * Shunt_Re + Shunt_Im * Shunt_Im)

        WattsDial_Re = A * (Re_R2R1 * alpha + Im_R2R1 * 0.0000033)
        WattsDial_Im = A * (Im_R2R1 * alpha - Re_R2R1 * 0.0000033)
        
        VarsDial_Re = B * (Beta * (CondC1 * Re_R2 - AdmtceC1 * Im_R2) + (AdmtceC1 * Re_R2 + CondC1 * Im_R2) * 0.000003)
        VarsDial_Im = B * (Beta * (AdmtceC1 * Re_R2 + CondC1 * Im_R2) - (CondC1 * Re_R2 - AdmtceC1 * Im_R2) * 0.000003)
        
        DetVolts_Re = -U * np.cos(DetPhse - DetAmpPhse) / (G * RefVolts)
        DetVolts_Im = -U * np.sin(DetPhse - DetAmpPhse) / (G * RefVolts)
        OffsetFactor_Re = 1 + Re_R2R1 + Re_R2 * Detcond - Im_R2 * DetAdmtce + Re_R2 * CondC1 - Im_R2 * AdmtceC1
        OffsetFactor_Im = Im_R2R1 + Re_R2 * DetAdmtce + Im_R2 * Detcond + Re_R2 * AdmtceC1 + Im_R2 * CondC1
        OffsetAdjustment_Re = DetVolts_Re * OffsetFactor_Re - DetVolts_Im * OffsetFactor_Im
        OffsetAdjustment_Im = DetVolts_Re * OffsetFactor_Im + DetVolts_Im * OffsetFactor_Re
        DialsPlusAdjust_Re = WattsDial_Re + VarsDial_Re + OffsetAdjustment_Re
        DialsPlusAdjust_Im = WattsDial_Im + VarsDial_Im + OffsetAdjustment_Im
        P = ScaleFactor_Re * DialsPlusAdjust_Re - ScaleFactor_Im * DialsPlusAdjust_Im
        Q = ScaleFactor_Re * DialsPlusAdjust_Im + ScaleFactor_Im * DialsPlusAdjust_Re

        AlphaCount = 1024 * P / ScaleFactor_Re

        BetaCount = 1024 * Q / (ScaleFactor_Re * (Re_R2 * AdmtceC1 + Im_R2 * CondC1))

        Volts = Divider_Re * RefVolts

        I_Re = (P * Divider_Re + Q * Divider_Im) / ((Divider_Re ** 2 + Divider_Im ** 2) * RefVolts)
        I_Im = (-P * Divider_Im + Q * Divider_Re) / ((Divider_Re ** 2 + Divider_Im ** 2) * RefVolts)
        amps = np.sqrt(I_Re ** 2 + I_Im ** 2)

        Watts = P
        Vars = Q

        Phase = self.phase(Q, P) * 180.0 / np.pi
        PowerFactor = np.cos(Phase)
        return {'Amps':amps,'Volts':Volts,'BetaCount':BetaCount,'AlphaCount':AlphaCount,'Watts':P,'Vars':Q,\
                'Phase':Phase,'PowerFactor':PowerFactor}
    
    def phase(self,y,x):
        if x == 0:
            return np.sign(y) * np.pi / 2.0
        theta = np.arctan(y / x)
        if x > 0:
            return theta
        return theta + np.sign(y) * np.pi
    
    def calculate(self,c,freq,set_volts,sheet,row):#As in the excel macros
        """I dont know what this code does, I copied the macros, and put all functions in one.
        Now will update the functions to use the complex maths in python, it will be much simpler.
        """
        creader = self.constants_reader
        
        Kraw = creader.get_Kraw(set_volts,sheet) #An array of three things, mean, uncertainty, DoF?
        #The excel sheet uses the uncal frequency, but i think it should be the calibrated frequency?
        Divider = Kraw[0] -1j*Kraw[0]*(freq['uncal']*Kraw[1]+Kraw[2])
        RefVolts = c[0]
        shunt = self.read_grid_cell(row,7) #not specifying a sheet defaults to the data sheet
        shunt_impedance = creader.get_shuntraw(shunt,sheet)
        Shunt = shunt_impedance[0]+1j*freq['uncal']*2*np.pi*shunt_impedance[1]*0.000001
        alpha = c[1]
        A = c[2]
        R2R1 = creader.get_r2_by_r1(sheet)+1j*creader.get_tc_r2_r1(sheet)*2*np.pi*freq['uncal']
        Beta = c[3]
        B = c[4] #Unused in new calculations?
        CondC1 = creader.get_g1(sheet)
        AdmtceC1 = creader.get_c1(sheet)*2*np.pi*0.000001*freq['uncal']
        R2 = creader.get_r2(freq['uncal'], sheet)
        U = c[5]
        DetPhse = c[6]
        G = c[7]
        DetAmpPhse = creader.get_det_amp_phase_gj(int(round(freq['uncal'],0)),sheet)
        Detcond = creader.get_gd(sheet)
        DetAdmtce = creader.get_yd(sheet)
        
        ScaleFactor = -Divider.conjugate()*RefVolts**2/Shunt
        WattsDial = A * (R2R1 * (alpha -1j*0.0000033))
        VarsDial = B*(Beta-0.000003*1j)*(CondC1+1j*AdmtceC1)*R2
        DetVolts = -U * np.exp(1j*(DetPhse - DetAmpPhse)) / (G * RefVolts)
        OffsetFactor = 1 + R2R1 + R2*(Detcond+1j*DetAdmtce)+R2*(CondC1+1j*AdmtceC1)
        OffsetAdjustment = DetVolts*OffsetFactor
        DialsPlusAdjust = WattsDial+VarsDial+OffsetAdjustment
        S = ScaleFactor*DialsPlusAdjust

        AlphaCount = 1024 * S.real / ScaleFactor.real
        BetaCount = 1024 * S.imag/(ScaleFactor.real*(R2.real*AdmtceC1 + R2.imag*CondC1))
        Volts = Divider.real * RefVolts
        I = S*Divider/RefVolts
        amps = np.abs(I)

        Phase =np.angle(S)
        PowerFactor = np.cos(Phase)
        return {'Amps':amps,'Volts':Volts,'BetaCount':BetaCount,'AlphaCount':AlphaCount,\
                'Watts':S.real,'Vars':S.imag,'Phase':Phase,'PowerFactor':PowerFactor}

    def load_dial_settings(self,settings):
        """Need to reread test point, it alsoe calcuates things and saves them somewhere to the sheet"""
        w_count,w_sign,v_count,v_sign=settings
        self.com(self.wbridge.set_value,'DV\r')
        self.com(self.wbridge.set_value,'W{}\r'.format(w_count))
        self.com(self.wbridge.set_value,'V\r'.format(v_count))
        self.com(self.wbridge.set_value,'{}\r'.format(w_sign))
        self.com(self.wbridge.set_value,'{}\r'.format(v_sign))

    def reading_loop(self,row):
        nordgs_col = 11
        N = int(self.read_grid_cell(row,nordgs_col))
        N = 1
        self.PrintSave("N changed to 1, for lazy reasons")
        for reading in range(N):
            self.com(self.wbridge.set_value,"DV\r\nA01\r\nB01\n")
            self.wait(0.5)
            gate_time = float(self.read_grid_cell(row,10))
            if gate_time>0.1: #if the ocunter is on
                #the original program has counter set ups for HP3131A too.
                self.com(self.count.set_value,"SENS:FREQ:GATE:TIME {}\n".format(gate_time))
                self.com(self.count.set_value,"SENS:FUNC 'FREQ 1'\n")
                self.com(self.count.set_value,"INIT\n")
                
            #GOD's AC
            time,Dcrms,Acrms,Err,Freq = self.run_swerl(row)
            
            # RUN FFT
            fft_ref_v,fft_ref_p = self.fft_func(row,Freq) #Is this the right frequency to send?
            self.com(self.wbridge.set_value,"DD\r\nA33\r\nB33\n")
            fft_det_v,fft_det_p = self.fft_func(row,Freq) #Is this the right frequency to send?
            
            #if the counter is on..
            if gate_time>0.1:
                self.com(self.count.set_value,"fetc?\n")
                self.wait(0.25)
                reading = self.com(self.count.read)
                counter_channels = raw_input("Both channels (0) or CH1 Only (1)?")
                if counter_channels == '0':
                    self.com(self.count.set_value,"SENS:FREQ:GATE:TIME {}\n".\
                             format(gate_time))
                    self.com(self.count.set_value,"SENS:FUNC 'FREQ 2'\n")
                    self.com(self.count.set_value,"read?\n")
                    self.wait(2)
                    reading = self.com(self.count.read)
                elif counter_channels == '1': #if not counter channels
                    watts_vars = str(self.read_grid_cell(row,13))
                    if watts_vars == 'vars':
                        #in testpoint it is asking if = to character 118 which is ASCII for v
                        pass
                        #call RD programs
                    elif watts_vars == 'watts': #if not watts or vars
                        pass
                        #call some other RD things
                    else:
                        self.PrintSave("Invalid watts or vars entry {}, NOT aborting".format(watts_vars))
                else:
                    self.PrintSave("Invalid counter channel '{}', not aborting but moving on".\
                                   format(counter_channels))
            
    def run_swerl(self, row):
        """
        start the swerlein algorithm, perhaps seperate thread?
        """
        self.PrintSave("Running GOD's AC at port {}".format(self.meter.bus))
        self.swerl = swerlein.Algorithm(self.inst_bus, port = self.meter.bus)
        #sets the thread running.
        loop = True
        #need to initialise the row as something,
        #then if program aborts mid measuremnt it wont crash on return.
        row = [0,0,0,0,0]
        #if results appear, row is overidden with actual data.
        while loop == True:
            if self.swerl.ready == True:
                #loop ends here and function returns the acdcrms readings
                row = self.swerl.All_data[0][2]
                self.set_grid_val(row,self.ref_v_print,acdcrms)
                loop = False
            elif self.swerl.error == True:
                #self._want_abort = 1
                self.PrintSave("Swerlein algorithm failed, NOT aborting")
                loop = False
            if self._want_abort:
                loop = False
        return row

    def fft_func(self,row,freq=50):
        """Function to read meter 256 times and fft the result.
        Returns the ninth elemnt of th etransformed frequency,
        this corresponds to the funcdamental, since we are measuring
        nine cycles of the wave form. It returns a complex number.
        Phase (currently the frequency of the ninth element? """
        #calculate the time to pause for.
        if freq ==0:
            freq = 50.0
            self.PrintSave("Changed frequency from 0 to 50")
        sample_time = 9.0/(256.0*freq)
        self.com(self.meter.set_value,"preset fast;mem fifo; mformat sint; oformat ascii\n")
        self.com(self.meter.set_value,"ssdc;range 10;ssrc ext\n")
        self.com(self.meter.set_value,"delay 1e-3;sweep {}, 256\n".format(sample_time))
        self.wait(2)
        series = np.zeros(256)
        self.PrintSave("Reading {} readings from {}".format(256,self.meter.label))
        self.printing_status = False#Stop logging info, or it will be just a mess
        for i in range(256):
            reading = float(self.com(self.meter.read))
            series[i] = reading
        self.printing_status = True #Start logging data again.
        fft_series = np.fft.rfft(series)
        #here should "sample_time" be used, or the calibrated (coorected) version?
        freq_vector = np.fft.rfftfreq(256,d=sample_time)
        self.PrintSave("Sample spacing of {} used, but those are uncalibrated,"+\
                       "should the calibrated spacing be used?".format(sample_time))
        magnitude = fft_series[9] #Magnitude of fundamnetal, as a complex number
        freq = freq_vector[9]*180/np.pi #frequency 
        return [magnitude,freq] #what goes on in here?

    def read_radian_all(self,row):
        pass

    def run(self,start_row = 12):
        """
        Main thread, reads through the table and executes commands.
        """
        row = start_row
        self.initialise_radian(row)
        new_rows = True #flag for existance of data at "row".
        while new_rows == True:
            self.set_power(row)
            self.power_sources(row)
            settings = self.find_dial_settings(row)
            self.reading_loop(row)
            self.read_radian_all(row)
            #The if statements could be replaced if this section of
            #code didnt need to know what source is plugged in.
            #Change to generic source?
            source_type = self.read_grid_cell(row,self.source_col)
            if source_type == "CH":
                ch_type = str( self.read_grid_cell(7,2) )
                if ch_type<99:# in:
                    self.com(self.ch5500.set_value,"S\n")
                    self.com(self.ch5050.set_value,"S\n")
            elif source_type == "HEG":
                self.com(self.heg.set_value,"SOUR:OPER:STOP\n")
            elif source_type =="FLUKE":
                self.fluke.shutdown()
            elif source_type == "FLUHIGH":
                self.fluke.shutdown()
            else:
                self.print_save("ERROR, no valid sources selected.")
            self.com(self.wbridge.set_value,"DV\rA01\rB01\r")
            self.wait(1)
            #self.paste_results?
            self.wait(1)

            #Measuremnt done, check if next row exists and movee on
            row = row+1
            if self.data_sh.cell(row=row,column = 2).value == None:
                new_rows = False
            if row>13:
                new_rows = False
        self.end()
        
if __name__ == "__main__":
    import visa2 #use simulated visa
    import fluke #The fluke object
    inst_bus = visa2

    #Create all the instruments as inst objects
    rd = inst.INSTRUMENT(inst_bus,label='RD',bus='GPIB::22::INSTR',SetValue='$')
    wbridge = inst.INSTRUMENT(inst_bus,label='WB',bus='ASRL1::INSTR',SetValue='$')
    #fl = inst.INSTRUMENT(visa2,label='FL',bus='ASRL1::INSTR')
    meter = inst.INSTRUMENT(inst_bus,label='HP',bus='GPIB::24::INSTR',SetValue='$')
    count = inst.INSTRUMENT(inst_bus,label = 'COUNT',bus = 'GPIB::11::INSTR',SetValue='$')
    ch5050 = inst.INSTRUMENT(inst_bus,label="CH5050",bus='GPIB::15::INSTR',SetValue='$')
    ch5500 = inst.INSTRUMENT(inst_bus,label="CH5500",bus='GPIB::15::INSTR',SetValue='$')
    heg = inst.INSTRUMENT(inst_bus,label="HEG",bus='GPIB::13::INSTR',SetValue='$')
    #except the fluke! Re use some old verified code
    fl = fluke.FLUKE(address='ASRL1::INSTR')

    #The workbook we will be using, contains all info imaginable. 
    wb = 'template2.xlsm'
    #The shared list, thread safe data structure.
    data = stuff.SharedList() #shared list thing
    EVT = None #Perhaps an EVT can be specified if a WX type GUI is used.
    notify_window = None #simulate a wx window?
    #A dictionary with the useful stuff inside.
    param = {'wb':wb,'rd':rd,'wbridge':wbridge,'meter':meter,'count':count,\
             'fluke':fl,'ch5500':ch5500,'ch5050':ch5050,'inst_bus':inst_bus,\
             'heg':heg}
    
    #And the thread is started here.
    thread = Thread(notify_window, EVT, param, data)

    
