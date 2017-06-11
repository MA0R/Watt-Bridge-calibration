"""
...words
"""
import stuff
import time
import numpy as np
import inst
import openpyxl
import threading
import swerlein

class Thread(threading.Thread):
    """
    Ref step main thread
    """
    def __init__(self, notify_window, EVT, param, data):
        threading.Thread.__init__(self, notify_window, EVT, param)
        self._want_abort = 0
        self.MadeSafe = False

        self.inst_bus = param['inst_bus']

        self.wb = param['wb']
        self.data_sh = wb['Data']

        self.rd = param['rd']
        self.rd.create_instrument()
        self.wbridge = param['wbridge']
        self.wbridge.create_instrument()
        self.meter = param['meter']
        self.meter.create_instrument()
        self.count = param['count']
        self.count.create_instrument()
        self.fluke = param['fluke']
        self.ch5500 = param['ch5500']
        self.ch5500.create_instrument()
        self.ch5050 = param['ch5050']
        self.ch5050.create_instrument()
        #need to get instruments from somewhere
        
        time_bit = time.strftime("%Y.%m.%d.%H.%M.%S, ",time.localtime())
        log_file_name = 'log.'+time_bit+".txt"
        self.raw_file_name = 'raw.'+time_bit
        self.logfile = open(log_file_name,'w')

        self.start()
        
    def abort(self):
        """abort worker thread."""
        # Method for use by main thread to signal an abort
        self._want_abort = 1
        
    def PrintSave(self, text):
        """
        Prints a string, and saves to the log file too.
        """
        if self._want_abort:
            #stop the pointless printing if the thread wants to abort
            return
        else:
            self.logfile.write(str(text)+"\n")
            print(str(text))
            
    def com(self, command,send_item=None):
        """
        Operate "command" on "send_item" if it is specified. Command is
        a method of the instrument class.
        All commands [Fluke? RD31?] go through this function, as it wraps each command
        with a check to see if the thread should abort.
        Recieve the status of the command from the instrument
        class, and if it failes this will flag the thread to abort.
        """
        if not self._want_abort:
            if send_item != None:
                sucess,val,string = command(send_item)
                if sucess == False:
                    self.abort()
                self.PrintSave(string)
                return val
            else:
                sucess,val,string = command()
                if sucess == False:
                    self.abort()
                self.PrintSave(string)
                return val
        else:
            return 0
        
    def MakeSafe(self):
        #probably not needed for internal use?
        pass

    def set_grid_val(self,row,col,data):
        """ Set the value of a grid cell at a row and column"""
        wx.CallAfter(self.grid.SetCellValue, row, col, str(data))
        
    def read_grid_cell(self,row,col,grid = None):
        """ Read a cell in an openpyxl worksheet object"""
        value = 0
        #make the data sheet the default grid to read from
        #could change and make grid be specified at each method call
        if grid == None:
            grid = self.data_sh
            
        if not self._want_abort:
            value = grid.cell(row=row, column=col).value
        return value
        
    def end(self):
        """
        clean up, turn instruments off [make safe?]
        """
        pass

    def wait(self,wait_time):
        """
        safely loop until time is up, checking for abort
        """
        self.PrintSave("Waiting for {} seconds".format(wait_time))
        t=time.time()
        while time.time()<t+wait_time:
            if self._want_abort: #has the stop button been pushed?
                break
        return

    def initialise_counter(self):
        pass

    def initialise_radian(self,row):
        """
        Initialisation process as in TP, with the phase sent
        to it after the wait
        """
        rd_phase_col = 16
        watts_or_vars_cel = 13
        
        self.wait(1)
        rd_phase = self.read_grid_cell(row,rd_phase_col)
        if rd_phase != None:
            self.com(self.rd.set_value,rd_phase)
        watts_or_vars = self.read_grid_cell(row,watts_or_vars_cel)
        if watts_or_vars != None:
            self.set_rd_output_pulse('prog_radian_id',1,1,rd_phase)
        else:
            pass
            #   :(
    def set_rd_output_pulse(self,a,b,c,d):
        pass
    
    def power_sources(self,row):
        source_col = 2
        
        source_type = self.read_grid_cell(row,source_col)
        if source_type == 'FLUKE':
            self.power_flue(row)
        elif source_type == 'FLUHIGH':
            self.power_fluke(row)
        elif source_type == 'CH':
            self.power_ch(row)
        elif source_type == 'HEG':
            self.power_heg(row)
        else:
            #abort since there is no valid source type
            self.PrintSave("Invalid source '{}' at row {}"\
                           .format(source_type,row))
            self._want_abort = 1

    def power_fluke(self,row):
        set_volts = self.read_grid_cell(row,4)
        set_phase = self.read_grid_cell(row,5)
        set_amps = self.read_grid_cell(row,3)
        set_freq = self.read_grid_cell(row,9)
        self.fluke.initialise()#can have fluke ramp time sent here
        #   AND WHAT ELSE?
        
        
    def power_ch(self,row):
        #repeated code? Already retrieved these cell values before.
        #could simplify
        set_volts = self.read_grid_cell(row,4)
        set_phase = self.read_grid_cell(row,5)
        set_amps = self.read_grid_cell(row,3)
        set_freq = self.read_grid_cell(row,9)
        #check first to see if ch5050 is wired in.
        if self.ch505:
            self.com(ch5050,"SI\n")
        self.com(self.ch5500.set_value,"S\n")
        if self.HVamp:
            self.com(self.ch5500.set_value,"O0\n")
            set_volts = self.amp_in_volts(set_volts)
        else:
            self.com(self.ch5500.set_value,"O180\n")
        self.com(self.ch5500.set_value,"R{}\n".format(set_volts))
        change_offset = True
        if change_offset:
            self.com(self.ch5500.set_value,"O0\n")
        self.com(self.ch5500.set_value,"V{}\n".format(np.abs(set_amps)))
        self.com(self.ch5500.set_value,"P{}F{}N\n".format(set_phase,set_freq))
        if self.ch5050:
            self.com(self.ch5050.set_value,"N\n")
            self.com(self.ch5500.set_value,"N\n")
            self.com(self.ch5050.set_value,"O\n")
        self.wait(60)
        
    def power_heg(self,row):
        pass

    def find_dial_settings(self):
        pass
    def refine_dial_settings(self):
        print("refining dial settings")

    def reading_loop(self,row):
        nordgs_col = 11
        N = int(self.read_grid_cell(row,nordgs_col))
        for reading in range(N):
            self.com(self.wbridge.set_value,"DV\r\nA01\r\nB01\n")
            self.wait(0.5)
            #the original program has counter set ups for HP3131A too.
            self.com(self.count.set_value,"SENS:FUNC 'FREQ 1'\n")
            self.com(self.count.set_value,"INIT\n")
            time,Dcrms,Acrms,Err,Freq = self.run_swerl(row)
            # SET UP FFT
            self.com(self.wbridge.set_value,"DD\r\nA33\r\nB33\n")
            # RUN FFT ?
            #if the counter is on..
            if 1==1:
                self.com(self.count.set_value,"fetc?\n")
                self.wait(0.25)
                reading = self.com(self.count.read_instrument)
                #if counter channels?
                if 1==1:
                    self.com(self.count.set_value,"SENS:FREQ:GATE:TIME\n")
                    self.com(self.count.set_value,"SENS:FUNC 'FREQ 2'\n")
                    #IF READING OR VARS
                    #SOMETHING TO DO WITH RD PROGRAMS?
            
    def run_swerl(self, row):
        """
        start the swerlein algorithm, perhaps seperate thread?
        """
        self.PrintSave("Running GOD's AC at port {}".format(self.meter.bus))
        self.swerl = swerlein.Algorithm(self.inst_bus, port = self.meter.bus)
        #sets the thread running.
        loop = True
        #need to initialise the row as something,
        #then if program aborts mid measuremnt it wont crash on return.
        row = [0,0,0,0,0]
        #if results appear, row is overidden with actual data.
        while loop == True:
            if self.swerl.ready == True:
                #loop ends here and function returns the acdcrms readings
                row = self.swerl.All_data[0][2]
                self.set_grid_val(row,self.ref_v_print,acdcrms)
                loop = False
            elif self.swerl.error == True:
                #self._want_abort = 1
                self.PrintSave("Swerlein algorithm failed, NOT aborting")
                loop = False
            if self._want_abort:
                loop = False
        return row

    def fft_func(row,freq=50):
        #calculate the time to pause for.
        sample_time = 9/(256*freq)
        self.com(self.meter.set_value,"preset fast;mem fifo; mformat sint; oformat ascii\n")
        self.com(self.meter.set_value,"ssdc;range 10;ssrc ext\n")
        self.com(self.meter.set_value,"delay 1e-3;sweep {}, 256\n".format(sample_time))
        self.wait(2)
        series = np.zeros(256)
        for i in range(256):
            reading = float(self.com(self.meter.read))
            series[i] = reading
        fft_series = np.fft.fft(series)

    def run(self,start_row = 12):
        """
        Main thread, reads through the table and executes commands.
        """
        
        row = start_row
        new_rows = True #flag for existance of data at "row".
        while new_rows == True:
            self.initialise_radian(row)
            self.power_sources(row)
            self.find_dial_settings()
            self.refine_dial_settings()
            self.reading_loop(row)
            

            row = row+1
            if self.data_sh.cell(row=row,column = 2).value == None:
                new_rows = False
            if row>13:
                new_rows = False
        self.end()
        
if __name__ == "__main__":
    import visa2
    import fluke
    inst_bus = visa2
    rd = inst.INSTRUMENT(inst_bus,label='RD',bus='GPIB::22::INSTR')
    wbridge = inst.INSTRUMENT(inst_bus,label='WB',bus='ASRL1::INSTR')
    #fl = inst.INSTRUMENT(visa2,label='FL',bus='ASRL1::INSTR')
    meter = inst.INSTRUMENT(inst_bus,label='FL',bus='GPIB::24::INSTR')
    count = inst.INSTRUMENT(inst_bus,label = 'COUNT',bus = 'GPIB::11::INSTR')
    fl = fluke.FLUKE(address='ASRL1::INSTR')
    ch5050 = inst.INSTRUMENT(inst_bus,label="CH5050",bus='GPIB::15::INSTR')
    ch5500 = inst.INSTRUMENT(inst_bus,label="CH5500",bus='GPIB::15::INSTR')
    
    wb = openpyxl.load_workbook('small_template.xlsx',data_only = False)
    data = stuff.SharedList() #shared list thing
    EVT = None
    notify_window = None #simulate a wx window?
    param = {'wb':wb,'rd':rd,'wbridge':wbridge,'meter':meter,'count':count,\
             'fluke':fl,'ch5500':ch5500,'ch5050':ch5050,'inst_bus':inst_bus}
    thread = Thread(notify_window, EVT, param, data)











    
