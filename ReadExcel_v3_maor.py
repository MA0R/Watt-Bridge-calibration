"""
This modified from "K2006vsWB GTC phase error_v1EDI_v3 5xpower_v4_noTrans.py"
(staff\TJS\GTC ex) for use with EDI K5.1 comparison.

This GTC program (using version 0.9.6) is used to read measurements from a
"Pb_Auto_V1xx.xls" based excel spreadsheet and emulate the WattBridge core
power calculation.  It prints error output and uncertainties to the consule
window for pasting back to the Table worksheet of the excel spreadsheet.
01 November 2013.
T J Stewart.
Further modified by KJ to avoid uncertain numbers not being available outside
function calls.

Version 3:  changed by TJS to split real and imaginary (or mag & phase) ureal
creation from combining as ucomplex.  This is to enable separate degrees of
freedom to be brought in for each part.

06/2017
MAOR BEN-SHAHAR
Modified to use openpyxl and so be compatible with .xlsx and .xlsm.
xlsm allows for macros, so it is used in the meantime but later
xlsx can be used instead.
This will modernise it a bit and make it compatible with other programs
already written with openpyxl (ref-step and lock-in. The lock in
can only work with openpyxl).
The changes involved referring to cells by numbers instead of letters,
and adding 1 to all rows and columns in EXCELDATA. Now col A  = col 1,
row 1 is the row labelled by 1 in excel too.
"""

import sys
import os
import numpy as np

progdir = os.path.dirname (os.path.abspath (sys.argv[0]))
sys.path.append(progdir)
#from core import * #this will be used as a module
#line above removed, -MB
import math
from GTC.library_complex import UncertainComplex
#Line above replaced with one below swapped, -MB
from GTC import ureal
from GTC import ucomplex
import string
#Line below removed, -MB
#import function
import xlrd
import openpyxl

class EXCELDATA(object):
    """
    Basic methods for reading data from worksheets
    """
    def __init__(self, workbook):
        self.workbook = workbook
        self.wb = openpyxl.load_workbook(self.workbook,data_only=True)
    def get_xl_rowdata(self, start_c, stop_c, select_r,sheet):
        row =[]
        for col in range(start_c,stop_c):
            row.append(sheet.cell(row=select_r,column=col).value)
        return row

    def get_xl_coldata(self, start_r, stop_r, select_c,sheet):
        col=[]
        for row in range(start_r,stop_r):
            col.append(sheet.cell(row=row,column=select_c).value)
        return col

    def get_xl_cell(self, colx, rowx, sheet):
        return sheet.cell(row=rowx,column=colx).value

class WATTDATA(EXCELDATA):
    """
    Specific locations of data for relevant watt bridge spreadsheet
    """
    def __init__(self, parent):
        EXCELDATA.__init__(self, parent)

    def get_buffer_gain(self, char, sheet):
        if char == '+':
            return self.get_xl_cell(4, 32, sheet) #sheet.cell_value(31,3)
        elif char == '-':
            return self.get_xl_cell(4, 33, sheet) # sheet.cell_value(32,3)
        else:
            return 100  #buffer gain must be '+' or '-' otherwise 100 will be obviously wrong

    def get_readings(self, start_c, rdata):
        dv = []
        for n in range(5):
            i = n*7 + start_c       # step 7 cells for next reading in row
            if rdata[i] !='':
                dv.append(rdata[i])
        return dv

    def get_Kraw(self, dR, sheet):
        if dR == 60:
            r = 39
        elif dR == 120:
            r = 40
        elif dR == 230:
            r = 41

        #return [sheet.cell_value(r,3), sheet.cell_value(r,4), sheet.cell_value(r,5)]
        return [self.get_xl_cell(4, r, sheet),self.get_xl_cell(5, r, sheet),self.get_xl_cell(6, r, sheet)]
        
    def get_shuntraw(self, sname, sheet):
        if sname == 'ESI(.1)':
            r = 22
        elif sname == 'C(.2)':
            r = 23
        elif sname == 'ESI(1)':
            r = 27
        elif sname == 'ESI(10)':
            r = 28
        elif sname == 'AuxDiv':
            print(" there was a mistake in the previous version, referring to cell 29 (which ends up being cell 30), this is a cell below the required cell.")
            r = 29
        #return [sheet.cell_value(r,3), sheet.cell_value(r,4)]
        return [self.get_xl_cell(4, r, sheet),self.get_xl_cell(5, r, sheet)]

    def get_DAmpPhse(self, f, sheet):
        f_r = round(f)
        flist = self.get_xl_coldata(46, 65, 3, sheet)
        phlist = self.get_xl_coldata(46, 65, 4, sheet)
        for i in range(len(flist)):
            if flist[i] == f_r:
                return phlist[i]
        return 0

    ## Definitions for equation variables (there are 13 in total for the Wattbridge equation)

    def get_Kstar(self, f, dRSet, sheet, uncL, dofL):
        Kraw = self.get_Kraw(dRSet, sheet)
        K_r = ureal(Kraw[0], Kraw[0]*uncL[0]/1000000, dofL[0], "Divider Ratio (K real)")
        K_i = ureal(Kraw[0]*(f*Kraw[1]+Kraw[2]), Kraw[0]*uncL[1]/1000000, dofL[1], "Divider Ratio (K imag)")
        return K_r, K_i
        ## K = complex(Kraw[0], Kraw[0]*(f*Kraw[1]+Kraw[2]))  # cells L10 and M10 of PowerCalc worksheet    
        ## return ucomplex(K.conjugate(), (K.real*uncL[0]/1000000, K.real*uncL[1]/1000000),dofL[0],"Divider Ratio (K*)")
        
    def get_HPcrctn(self, sheet, uncL, dofL):
        HP3458corectn = self.get_xl_cell(4, 7, sheet)
        HP3458corectn_u = HP3458corectn*uncL[2]/1000000 # 3458 type b uncertainty for GodsAC included here
        HP_crctn = ureal(HP3458corectn, HP3458corectn_u, dofL[3], "Divider Output Voltage (V1)")
        return HP_crctn

    def get_ShuntByTrans(self, f, sName, sheet, transR, uncL, dofL):
        shuntraw = self.get_shuntraw(sName, sheet)
        shunt_r = ureal(shuntraw[0], shuntraw[0]*uncL[3]/1000000, dofL[3], "Shunt Resistance (R3 real)")
        shunt_i = ureal(f*2*math.pi*shuntraw[1]*0.000001, shuntraw[0]*uncL[4]/1000000, dofL[4], "Shunt Resistance (R3 imag)")
        
        ## shnt = complex(shuntraw[0], f*2*math.pi*shuntraw[1]*0.000001)  # cells O10 and P10 of PowerCalc worksheet
        ## Shunt = ucomplex(shnt,(shnt.real*uncL[3]/1000000, shnt.real*uncL[4]/1000000), dofL[3], "Shunt Resistance (R3)")
        Trans = ucomplex(complex(transR, 0), (transR*2/1000000, 0), 6, "Shunt by Transformer Resistance (R3)") ######## ! zeros as placeholder for phase & uncertainty of Transformer ratio
        if transR >1:
            return Shunt/Trans
        else:
            return shunt_r, shunt_i
        
    def get_alpha(self, wd, v1, uncL, dofL):
        alpha_r = ureal(wd, (wd*uncL[5]+uncL[21])/1000000, dofL[5], "Alpha Ratio (alpha real)")
        alpha_i = ureal(-v1*0.0000033, wd*uncL[23]/1000000+5.7e-7, dofL[23], "Alpha Ratio (alpha imag)")
        ## alph = wd-v1*0.0000033j     # real part is cell Q10 of PowerCalc, imag is from offset in WattFuncs calculate() macro.  Also include offset as being proportional to V1.
        ## alpha_u = ((wd*uncL[5]+uncL[21])/1000000, wd*1/1000000+5.7e-7) #imag uncertainty is 1 ppm of dial setting plus offset uncertainty from cell BJ22 of "IVD offset SR830 readings May12.xls"
        ## return ucomplex(alph, alpha_u, dofL[5], "Alpha Ratio (alpha)")
        return alpha_r, alpha_i

    def get_A(self, ws, sheet, uncL, dofL):
        A_1 = self.get_buffer_gain(ws, sheet) 
        return ureal(A_1, abs(A_1*uncL[6]/1000000), dofL[6], "Alpha Buffer Gain (A)")    # cell R10 of PowerCalc
        
    def get_R1R2(self, f, sheet, uncL, dofL):
        R2R1r = self.get_xl_cell(4, 12, sheet)     # cell S10 of PowerCalc
        R2R1i = self.get_xl_cell(4, 13, sheet)*f*2*math.pi      # cell T10 of PowerCalc  
        R2R1_r = ureal(R2R1r, abs(R2R1r*uncL[7]/1000000), dofL[7], "R2 R1 Ratio (R2/R1 real)")
        R2R1_i = ureal(R2R1i, abs(R2R1r*uncL[8]/1000000), dofL[8], "R2 R1 Ratio (R2/R1 imag)")
        ## return ucomplex(complex(R2R1_re, R2R1_im), (abs(R2R1_re*uncL[7]/1000000), abs(R2R1_re*uncL[8]/1000000)), dofL[7], "R2 R1 Ratio (R2/R1)")
        return R2R1_r, R2R1_i
        
    def get_beta(self, vd, v1, uncL, dofL):
        beta_r = ureal(vd, (vd*uncL[9]+uncL[22])/1000000, dofL[9], "Beta Ratio (beta real)")
        beta_i = ureal(-v1*0.0000030, vd*uncL[24]/1000000+3.5e-7, dofL[24], "Beta Ratio (beta imag)")
        ## bet = vd-v1*0.0000030j      # real part is cell U10 of PowerCalc, imag is from offset in WattFuncs calculate() macro.  Also include offset as being proportional to V1.
        ## beta_u = ((vd*uncL[9]+uncL[22])/1000000, vd*1/1000000+3.5e-7) #imag uncertainty is 1 ppm of dial setting plus offset uncertainty from cell BI22 of "IVD offset SR830 readings May12.xls"
        ## return ucomplex(bet, beta_u, dofL[9], "Beta Ratio (beta)")
        return beta_r, beta_i

    def get_B(self, vs, sheet, uncL, dofL):
        B_1 = self.get_buffer_gain(vs, sheet)
        return ureal(B_1, abs(B_1*uncL[10]/1000000), dofL[10], "Beta Buffer Gain (B)")    # cell V10 of PowerCalc

    def get_Y1(self, f, sheet, C1Tcrct,uncL, dofL):
        Y1r = self.get_xl_cell(4, 15, sheet)     # cell W10 of PowerCalc
        Y1i = f*2*math.pi*C1Tcrct*0.000001    # cell X10 of PowerCalc
        Y1_r = ureal(Y1r, Y1i*uncL[11]/1000000, dofL[11], "Capacitor Admittance (Y1 real)")
        Y1_i = ureal(Y1i, Y1i*uncL[12]/1000000, dofL[12], "Capacitor Admittance (Y1 imag)")
        ## g_1 = self.get_xl_cell('D', 15, sheet)     # cell W10 of PowerCalc
        ## Y_1 = f*2*math.pi*C1Tcrct*0.000001    # cell X10 of PowerCalc
        ## C1_u = (Y_1*uncL[11]/1000000, Y_1*uncL[12]/1000000) # """!!! change from orignal excel template !!!""" 
        ## return ucomplex(complex(g_1, Y_1), C1_u, dofL[12], "Capacitor Admittance (Y1)")
        return Y1_r, Y1_i
        
    def get_R2(self, f, sheet, uncL, dofL):
        R2r = self.get_xl_cell(4, 10, sheet)
        R2i = self.get_xl_cell(4, 11, sheet)*f*2*math.pi
        R2_r = ureal(R2r, R2r*uncL[13]/1000000, dofL[13], "Resistor (R2 real)")
        R2_i = ureal(R2i, R2r*uncL[14]/1000000/(2*f*math.pi), dofL[14], "Resistor (R2 imag)")
        ## R_2 = complex(self.get_xl_cell('D', 10, sheet), self.get_xl_cell('D', 11, sheet)*f*2*math.pi)  # cells Y10 and Z10 of PowerCalc 
        ## R2_u = (R_2.real*uncL[13]/1000000, 10000*uncL[14]/1000000/(2*f*math.pi))
        ## return ucomplex(R_2, R2_u, dofL[13], "Resistor (R2)")
        return R2_r, R2_i
        
    def get_R1(self, f, sheet, uncL, dofL):
        R1r = self.get_xl_cell(4, 8, sheet)
        R1i = self.get_xl_cell(4, 9, sheet)*f*2*math.pi
        R1_r = ureal(R1r, R1r*uncL[13]/1000000, dofL[13], "Resistor (R1 real)")  # use same uncertainty as for R2
        R1_i = ureal(R1i, R1r*uncL[14]/1000000/(2*f*math.pi), dofL[14], "Resistor (R1 imag)")
        return R1_r, R1_i
        
    def get_dVu_b(self, uncL, dofL):
        return ureal(1, uncL[15]/1000000, dofL[15], "Detector Voltage (U mag)")   # get Detector volts type B unc (in ppm)
        
    def get_dPhu_b(self, uncL, dofL):
        return ureal(0, uncL[16]/1000000, dofL[16], "Detector Voltage Phase (U phase)") # get Detector phase type B unc (in micro radians)
        
    def get_rPhu_b(self, uncL, dofL):
        return ureal(0, uncL[16]/1000000, dofL[16], "V1 Reference Phase (V1 phase)") # get reference phase type B unc (in micro radians)
        
    def make_DetUmean(self, dv, dp, rp, dvub, dpub, rpub):
        dV = [i*dvub for i in dv]                      # create list of ureals from the detector volts readings and type B unc
        dPh = [math.radians(i)+dpub for i in dp]      # create list of ureals from the gross detector phase readings and type B unc
        rPh = [math.radians(i)+rpub for i in rp]      # create list of ureals from the reference phase readings and type B unc
        dPhs = []
        for i in range(len(dPh)):
            dPhs.append(dPh[i]-rPh[i])                  # create list of detector phase readings wrt reference phase

        DetU = []
        for i in range(len(dPhs)):
            DetU.append(UncertainComplex(dV[i]*math.cos(dPhs[i]), dV[i]*math.sin(dPhs[i])))   # put magnitude and phase into real & imag parts
        #print DetU
        return DetU
               
    def get_DetGain(self, f, sheet, uncL, dofL):
        DetG = self.get_xl_cell(4, 20, sheet)        # cell AC10 of PowerCalc
        DetAPh = math.radians(self.get_DAmpPhse(f, sheet))   # cell AD10 of PowerCalc 
        G_mag = ureal(DetG, DetG*uncL[17]/1000000, dofL[17], "Detector Gain (G mag)")
        G_ph = ureal(DetAPh, uncL[18]/1000000, dofL[18], "Detector Gain Phase (G phase)")
        #return UncertainComplex(DetAmpGain*cos(DetAmpPhse), DetAmpGain*sin(DetAmpPhse))
        return G_mag, G_ph

    def get_DetC(self, f, sheet, uncL, dofL):
        DCr = self.get_xl_cell(4, 17, sheet)
        DCi = self.get_xl_cell(4, 16, sheet)*f*2*math.pi*1e-12  # cells AE10 and AF10 of PowerCalc
        ## DetC_u = (Det_C.real*uncL[19]/1000000,Det_C.imag*uncL[20]/1000000)
        DetC_r = ureal(DCr, DCr*uncL[19]/1000000, dofL[19], "Detector Admittance (Yd real)")
        DetC_i = ureal(DCi, DCi*uncL[20]/1000000, dofL[20], "Detector Admittance (Yd imag)")
        ## return ucomplex(Det_C, DetC_u, dofL[20], "Detector Admittance (Yd)")
        return DetC_r, DetC_i
##  Definitions for DUC

    def get_DUC_phasedata(self, start_c, stop_c, rdata):
        d = rdata[start_c:stop_c]
        for i in range(len(d)):       #K2006 reports phase angle from 0 - 360, convert to -180 to 180 degrees.
            if d[i] > 180:
                d[i]-=360
        return d

    def get_DUC_mean(self, data, TBu, TBdof):
        uB = ureal(0, TBu, TBdof, "DUC_typeB")
        d = [i+uB for i in data]
        d_mean1 = sum(d)/len(d)
        d_mean2 = type_a.estimate([float(di) for di in d],"DUCmean_typeA")
        return d_mean1 - float(d_mean1) + d_mean2
    
####################################
##  ADDITIONAL FUNCTIONS BY MAOR  ##
####################################
    def get_watts_buffer_gain(self,sign,sheet):
        col = 4
        if sign == 'WP+':
            r=32
        elif sign == 'WP-':
            r=33
        elif sign == 'VP+':
            r=34
        elif sign == 'VP-':
            r=35
        else:
            r=35
            pass #raise some error
        return float(self.get_xl_cell(col, r, sheet))
    def get_r2_by_r1(self,sheet):
        return float(self.get_xl_cell(4,12,sheet))
    def get_tc_r2_r1(self,sheet):
        return float(self.get_xl_cell(4,13,sheet))
    def get_g1(self,sheet):
        return float(self.get_xl_cell(4,15,sheet))
    def get_c1(self,sheet):
        return float(self.get_xl_cell(4,14,sheet))
    def get_r1(self,f,sheet):
        return complex(self.get_xl_cell(4,8,sheet),2*np.pi*f*self.get_xl_cell(4,9,sheet))
    def get_r2(self,f,sheet):
        return complex(self.get_xl_cell(4,10,sheet),2*np.pi*f*self.get_xl_cell(4,11,sheet))
    def get_amp_gain(self,val,sheet):
        if val == 'Min':
            r = 19
        elif val == 'Max':
            r = 20
        else:
            r=None
        return float(self.get_xl_cell(4,r,sheet))
    def get_det_amp_phase_gj(self,round_f,sheet):
        cells = self.get_xl_coldata(46,64,3,sheet)
        cells = [int(c) for c in cells] #need to convert from Long to int
        consts = self.get_xl_coldata(46,64,4,sheet)
        idx = cells.index(round_f)
        constant = consts[idx]
        ### But there would never be a case of 49.5??
        return float(constant)
    def get_gd(self,sheet):
        return float(self.get_xl_cell(4,17,sheet))
    def get_yd(self,sheet):
        return float(self.get_xl_cell(4,16,sheet)*0.000000000001)
    
############################################

    def phase_error_summary(self, typ, rdata, phi):
        ducTBu = rdata[13]
        ducTBdof = rdata[14]        
        ## Depending on type, get the phase data from DUC, determine their uncertain numbers and return an error string
        if typ == 'UtoI':   # three channels to calculate error for
            DUC_ch1 = self.get_DUC_phasedata(160, 165, rdata)
            DUC_ch2 = self.get_DUC_phasedata(165, 170, rdata)
            DUC_ch3 = self.get_DUC_phasedata(170, 175, rdata)
            DUC_1mean  = self.get_DUC_mean(DUC_ch1, ducTBu, ducTBdof)
            DUC_2mean  = self.get_DUC_mean(DUC_ch2, ducTBu, ducTBdof)
            DUC_3mean  = self.get_DUC_mean(DUC_ch3, ducTBu, ducTBdof)
            phi_E1 = DUC_1mean-phi
            phi_E2 = DUC_2mean-phi
            phi_E3 = DUC_3mean-phi
            return "%G\t%G\t%G\t%G\t%G\t%G\t%G\t%G\t%G" %(phi_E1.x, phi_E1.u, phi_E1.df,phi_E2.x, phi_E2.u, phi_E2.df, phi_E3.x, phi_E3.u, phi_E3.df)
        else:
            DUC_ch1 = self.get_DUC_phasedata(160, 165, rdata)
            DUC_1mean  = self.get_DUC_mean(DUC_ch1, ducTBu, ducTBdof)
            phi_E1 = DUC_1mean-phi    
            return "%G\t%G\t%G" %(phi_E1.x, phi_E1.u, phi_E1.df)

if __name__ == "__main__" :
    print 'Testing components'
    excel = WATTDATA("APMP-K5.1 mod_v2_Calcs _d.xlsm")
    shData = excel.wb['Data']
    sheetC = excel.wb['constants']
    sheetU = excel.wb['Uncertainties']
    shTable = excel.wb['Excel Results']
    uncList = excel.get_xl_coldata(5,27,3,sheetU)
    dofList = excel.get_xl_coldata(5,27,5,sheetU)
    #uncList = sheetU.col_values(2,4,27)  # get uncertainty list from Uncertainties worksheet
    #dofList = sheetU.col_values(4,4,27)  # get degrees of freedom list from Uncertainties worksheet
    print repr(excel.get_Kstar(50,230,sheetC, uncList, dofList))

    ## rdata = excel.get_xl_rowdata(0,177,269,shData)
    ## dVolts = excel.get_readings(32, rdata)
    ## dPhase = excel.get_readings(33, rdata)
    ## refVolts = excel.get_readings(34, rdata)
    ## refPhase = excel.get_readings(36, rdata)   
    ## det_b = excel.get_Det_b(uncList, dofList)
    ## det_Umean = excel.make_DetUmean(dVolts, dPhase, refPhase, det_b[0], det_b[1], det_b[2])
    
